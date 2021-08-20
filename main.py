import requests
from bs4 import BeautifulSoup
import os
from openpyxl import Workbook, load_workbook
from flask import Flask, render_template, request, send_file, redirect
import sqlite3
import datetime
import time
from fake_useragent import UserAgent

app = Flask(__name__)


def db_writer_error(category_name, url, error):
    conn = sqlite3.connect('logs.db')
    cursor = conn.cursor()
    if type(error) == str and error == "":
        cursor.execute("insert into BURL values (?, ?)", (category_name, url))
    elif type(error) == list:
        cursor.execute("insert into PLP values (?, ?, ?, ? ,? ,? ,? ,? ,? ,?, ?, ?)", (
            error[0], error[1], error[2], error[3], error[4], error[5], error[6], error[7], error[8], error[9], error[10], error[11]))
    else:
        cursor.execute("insert into ERRORS values (?, ?, ?)",
                       (category_name, url, error))
    conn.commit()
    conn.close()


def crawler(url, type):
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"}
    response = requests.get(url, headers=headers)
    print(url,type)
    # time.sleep(10)
    soup = BeautifulSoup(response.text, 'lxml')

    browse_version = soup.find_all(
        'div', attrs={'class': 'product-wrapper row--hasPopup'})
    # Base Case - Error 404
    error404 = soup.find(
        'div', attrs={'class': 'col-xs-12 col-sm-11 col-sm-offset-1'})
    if error404 is not None and error404.find('strong').text == "Ooops! Page not found.":
        return "Error404"

    # Base Case - None Page
    try:
        filterby = soup.find('aside', attrs={'class': 'col-sm-3 left-rail'})
        filterbyfind = filterby.find('h2').text
        nonepage = soup.find_all(
            'div', attrs={'class': 'col-sm-3 subcat-thumb'})
        if nonepage == [] and browse_version == [] and filterbyfind == "Filter By":
            return "NonePage"
    except:
        pass

    # Base Case Image Error 404
    try:
        if type == "type4":
            server_error =  soup.find('div', attrs={'id':'header'})
            if server_error.find('h1').text == "Server Error":
                return "Error404"
    except:
        pass
    
    # Base Case - Browse Version

    if browse_version:
        header = soup.find('header', attrs={'class': 'col-xs-12 page-header'})
        category_name = header.find('h1').text

        if type == "type2":
            nrpp = soup.find(
                'span', attrs={'class': 'search-results-amt'}).text.split(' ')[0][1:]
            new_url = url.split('?')[0] + "?No=0&Nrpp=" + str(nrpp)
            db_writer_error(category_name, new_url, "")

        elif type == "type3":
            for container in browse_version:
                pl = container.find('div', attrs={'class': 'image'})
                pl_url = "https://www.grandandtoy.com"+pl.find('a')['href']
                pl_image_url = pl.find('img')['src']
                pl_title = pl.find('img')['alt']
                availabilityPLP = container.find(
                    'div', attrs={'class': 'availabilityMsg'})
                availabilityPLP = availabilityPLP.find('p').text

                imageres = crawler(pl_image_url, "type4")
                if imageres is None:
                    broken_imagePLP = "No"
                else:
                    broken_imagePLP = "Yes"
                res = crawler(pl_url, type)
                if res is None:
                    continue
                elif res == "Error404":
                    db_writer_error("", "", [
                                    pl_url, pl_title, res, broken_imagePLP, availabilityPLP, "NA", "NA", "NA", "NA", "NA","NA", url])
                elif res == "NonePage":
                    db_writer_error("", "", [
                                    pl_url, pl_title, res, broken_imagePLP, availabilityPLP, "NA", "NA", "NA", "NA", "NA","NA", url])
                else:
                    # print(res)
                    db_writer_error(
                        "", "", [pl_url, pl_title, "No", broken_imagePLP, availabilityPLP]+res+[url])
            return None

        elif type == "type1":
            # Find Category in Filter By
            filter_container = soup.find(
                'div', attrs={'class': 'filters-container'})
            filters = filter_container.find_all(
                'h3', attrs={'class': 'js-accordion__header'})
            for filter in filters:
                if filter.text == 'Category':
                    db_writer_error(category_name, url,
                                    "Category page appearing as Browse")
                    return None
    if type == "type4":
        return None

    # Category Iteration
    categories = soup.find_all('div', attrs={'class': 'col-sm-3 subcat-thumb'})
    for category in categories:
        url = "https://www.grandandtoy.com"+category.find('a')['href']
        category_name = category.find('span').text
        image_url = "https://www.grandandtoy.com"+category.find('img')['src']
        if type == "type3":
            return None
        res = crawler(url, type)

        if type == "type1":
            if res == "Error404":
                db_writer_error(category_name, url, "Page Error 404")
            elif res == "NonePage":
                db_writer_error(category_name, url, "NonePage")

            res = crawler(image_url, "type1")
            if res == "Error404":
                db_writer_error(category_name, url, "Image Error 404")

    if type == "type3":
        availabilityPDP = soup.find('div', attrs={'class': 'availabilityMsg'})
        availabilityPDP = availabilityPDP.find('p').text
        pdp_title_header = soup.find('div', attrs={'class': 'product-header'})
        pdp_title = pdp_title_header.find('h1')
        if pdp_title is not None and len(pdp_title.text) > 0:
            title_disp = "Yes"
        else:
            title_disp = "No"
        
        char_count = 0
        parent = soup.find('div', attrs={'class':'js-accordion__panel product-subDetails'})
        if parent is not None:
            children = parent.findChildren(recursive=True)
            # print(parent.get_text())

            # with open("temp.html", 'w') as f:
            #     f.write(str(response.text))
            for child in children:
                try:
                    if child.text == "SPECIFICATIONS":
                        break
                    if child.text == "DETAILS":
                        continue
                    char_count += len(child.text)
                except:
                    pass
        else:
            char_count = 0

        if char_count >= 500:
            count_details = "Yes"
        else:
            count_details = "No"
        
        if char_count == 0:
            details_showing = "No"
        else:
            details_showing = "Yes"
        
        small_image = soup.find('ul', attrs={'class':'easyzoom-thumbnails thumbnails'})
        if small_image is not None:
            image_count = len(small_image.find_all('li'))
            if image_count>=3:
                multiple_images = "Yes"
            else:
                multiple_images = "No"
        else:
            multiple_images = "No"

        pd_image_div = soup.find('div', attrs={'id':'standard-image'})
        if pd_image_div is not None:
            pd_image_url = pd_image_div.find('img')['src']
            if pd_image_url[:-3] == "jpg":
                if pd_image_url[:8] != "https://":
                    pd_image_url = "https://www.grandandtoy.com" + pd_image_url
                    imageres = crawler(pd_image_url, "type4")
                else:
                    imageres = crawler(pd_image_url, "type4")
            else:
                imageres = None
            if imageres is None:
                broken_imagePDP = "No"
            else:
                broken_imagePDP = "Yes"
        else:
            broken_imagePDP = "Yes"
        return [broken_imagePDP, availabilityPDP, multiple_images, title_disp, count_details, details_showing]

    if type == "type1":
        category_name = url.split("/")[-3]
        # Banner Checking Type 1
        bannert1 = soup.find(
            'a', attrs={'class': 'btn btn-primary mainBanner'})
        if bannert1 is not None:
            bannert1_url = bannert1['href']
            if bannert1_url[:8] != "https://":
                bannert1_url = "https://www.grandandtoy.com"+bannert1_url
                res = crawler(bannert1_url, "type1")
            else:
                res = crawler(bannert1_url, "type1")

            if res == "Error404":
                db_writer_error(category_name, url, "Banner Page Error 404")
            elif res == "NonePage":
                db_writer_error(category_name, url, "Banner None Page")

            image_banner_t1 = soup.find(
                'img', attrs={'class': 'col-md-6 hidden-xs p-0'})
            if image_banner_t1 is not None:
                image_banner_t1_url = image_banner_t1['src']
                res = crawler(image_banner_t1_url, "type1")
                if res == "Error404":
                    db_writer_error(category_name, url,
                                    "Banner Image Error 404")

        # Banner Check Type 2
        bannert2 = soup.find(
            'div', attrs={'class': 'cmain-top row row--padded'})
        if bannert2 is not None and bannert1 is None:
            bannert2_url = bannert2.find('a')['href']
            if bannert2_url[:8] != "https://":
                bannert2_url = "https://www.grandandtoy.com" + bannert2_url
                res = crawler(bannert2_url, "type1")
            else:
                res = crawler(bannert2_url, "type1")
            if res == "Error404":
                db_writer_error(category_name, url, "Banner Page Error 404")
            elif res == "NonePage":
                db_writer_error(category_name, url, "Banner None Page")

            image_banner_t2 = soup.find(
                'img', attrs={'class': 'col-md-12 hidden-xs p-0'})
            if image_banner_t2 is not None:
                image_banner_t2_url = image_banner_t2['src']
                res = crawler(image_banner_t2_url, "type1")
                if res == "Error404":
                    db_writer_error(category_name, url,
                                    "Banner Image Error 404")
    return None


@app.route("/", methods=["GET", "POST"])
def home():
    if request.method == "POST" and request.form.get("finderrors"):
        url = request.form.get("url")
        if url != "":
            conn = sqlite3.connect('logs.db')
            cursor = conn.cursor()
            cursor.execute("DROP TABLE IF EXISTS ERRORS")
            sql = '''CREATE TABLE ERRORS(
            CategoryName CHAR(20) NOT NULL,
            Url CHAR(20) NOT NULL,
            Comments CHAR(20) NOT NULL
            )'''
            cursor.execute(sql)
            conn.commit()

            excel_dir = os.path.expanduser("~")
            wb = Workbook()
            ws = wb.active
            ws.cell(row=1, column=1).value = "Category Name"
            ws.cell(row=1, column=2).value = "Link"
            ws.cell(row=1, column=3).value = "Comments"

            try:
                res = crawler(url, "type1")
                url_name = url.split("/")[-3]
                if res == "Error404":
                    db_writer_error(url_name, url, "Page Error 404")
                elif res == "NonePage":
                    db_writer_error(url_name, url, "None Page")
            except Exception as e:
                print(e)

            row = 2
            for rows in cursor.execute('SELECT * FROM ERRORS'):
                for col in range(1,4):
                    ws.cell(row=row, column=col).value = rows[col-1]
                row += 1
            wb.save(excel_dir + "\\temp.xlsx")
            wb.close()
            print("Finished Type 1")
            conn.close()
            return render_template("newhome.html", condition1=True, condition2=False)

    elif request.method == "POST" and request.form.get("download_fe"):
        return redirect('/download_fe')

    elif request.method == "POST" and request.form.get("findplp"):
        url = request.form.get("url")
        if url != "":
            conn = sqlite3.connect('logs.db')
            cursor = conn.cursor()
            cursor.execute("DROP TABLE IF EXISTS BURL")
            sql = '''CREATE TABLE BURL(
            CategoryName CHAR(20) NOT NULL,
            Url CHAR(20) NOT NULL
            )'''
            cursor.execute(sql)
            conn.commit()

            excel_dir = os.path.expanduser("~")
            wb = Workbook()
            ws = wb.active
            ws.cell(row=1, column=1).value = "Category Name"
            ws.cell(row=1, column=2).value = "Link"

            try:
                res = crawler(url, "type2")
                url_name = url.split("/")[-3]
                if res == "Error404":
                    db_writer_error(url_name, url, "")
                elif res == "NonePage":
                    db_writer_error(url_name, url, "")
            except Exception as e:
                print(e)
            row = 2
            for rows in cursor.execute('SELECT * FROM BURL'):
                ws.cell(row=row, column=1).value = rows[0]
                ws.cell(row=row, column=2).value = rows[1]
                row += 1
            wb.save(excel_dir + "\\burl.xlsx")
            wb.close()
            print("Finished Type 2")

            cursor.execute("DROP TABLE IF EXISTS PLP")
            sql = '''CREATE TABLE PLP(
            ProductURL CHAR(20) NOT NULL,
            ProductTITLE CHAR(20) NOT NULL,
            ErrorPage CHAR(20) NOT NULL,
            BrokenImagePLP CHAR(20) NOT NULL,
            AvaiabilityPLP CHAR(20) NOT NULL,
            BrokenImagePDP CHAR(20) NOT NULL,
            AvaiabilityPDP CHAR(20) NOT NULL,
            MultipleImages CHAR(20) NOT NULL,
            TitleDisplayed CHAR(20) NOT NULL,
            CharCount CHAR(20) NOT NULL,
            DetailsDisp CHAR(20) NOT NULL,
            BrowseURL CHAR(20) NOT NULL
            )'''
            cursor.execute(sql)
            conn.commit()
            

            wb = Workbook()
            ws = wb.active
            ws.cell(row=1, column=1).value = "Product URL"
            ws.cell(row=1, column=2).value = "Product Title"
            ws.cell(row=1, column=3).value = "Leading Error Page"
            ws.cell(row=1, column=4).value = "Broken Image PLP"
            ws.cell(row=1, column=5).value = "Availability PLP"
            ws.cell(row=1, column=6).value = "Broken Image PDP"
            ws.cell(row=1, column=7).value = "Availability PDP"
            ws.cell(row=1, column=8).value = "Multiple Images"
            ws.cell(row=1, column=9).value = "Title Displayed"
            ws.cell(row=1, column=10).value = "Char count of details>=500"
            ws.cell(row=1, column=11).value = "Details showing"
            ws.cell(row=1, column=12).value = "Browse URL"

            work_book = load_workbook(
                excel_dir + "\\burl.xlsx", read_only=False)
            work_sheet = work_book.active

            for row in range(2, work_sheet.max_row + 1):
                res = crawler(work_sheet.cell(row=row, column=2).value, "type3")
            row = 2
            for rows in cursor.execute('SELECT * FROM PLP'):
                for col in range(1,13):
                    ws.cell(row=row, column=col).value = rows[col-1]
                row += 1


            wb.save(excel_dir + "\\plp.xlsx")
            wb.close()
            conn.close()
            return render_template("newhome.html", condition1=False, condition2=True)

    elif request.method == "POST" and request.form.get("download_plp"):
        return redirect('/download_plp')

    return render_template("newhome.html", condition1=False, condition2=False)


@app.route("/download_fe", methods=["GET", "POST"])
def download_fe():
    path = os.path.expanduser("~") + "\\temp.xlsx"
    return send_file(path, as_attachment=True)


@app.route("/download_plp", methods=["GET", "POST"])
def download_plp():
    path = os.path.expanduser("~") + "\\plp.xlsx"
    return send_file(path, as_attachment=True)


if __name__ == "__main__":
    app.run(debug=True)
    # t1 = datetime.datetime.now()
    # ua = UserAgent()
    # res = requests.get("https://www.grandandtoy.com/EN/search/pages/browse/Office-Supplies/_/N-af2?No=0&Nrpp=9001")
    # soup = BeautifulSoup(res.text, 'html.parser')
    # with open("test.html" , "w" , encoding="utf-8") as f:
    #     f.write(str(soup))
    # t2 = datetime.datetime.now()
    # print(t2-t1)
